__author__ = 'Caitrin'
# encoding: utf-8

import os
import itertools
from openpyxl import load_workbook
#import cPickle as pickle
import xml.etree.ElementTree as ET
from preprocessingCodeLang import Preprocessor
import ast
from multiprocessing import Pool
import pathos.pools as pp
import shutil
import enchant
class BugReport:

    def __init__(self, reportID=None, bug_id=None, summary=None, description=None, report_time=None, report_timestamp=None, status=None, commit=None, commit_timestamp=None, files=None, filesLong=None):
        self.reportID = reportID
        self.bug_id = bug_id
        self.summary = summary
        self.description = description
        self.report_time = report_time
        self.report_timestamp = report_timestamp
        self.status = status
        self.commit = commit
        self.commit_timestamp = commit_timestamp
        self.files = files
        self.filesLong = filesLong
        self.processed_description = ""

class DataProcessor:

    def __init__(self):
        pass
    def get_stackoverflow_data_sentences_all(self, dirs):
        sent = []
        d = enchant.Dict("en_US")
        count = 0

        for dir in dirs:
            for filename in os.listdir(dir):
                count += 1
               # print count
                with open(dir + filename, 'r') as content_file:
                    for line in content_file:
                        tokens = line.strip().split(",")
                        if not tokens:
                            continue
                        code = [s for s in tokens if "@" in s]

                        nl = [s for s in tokens if "@" not in s]

                        sent.append(tokens)

                        for t in code:
                            if d.check(t[1:-1]):
                                sent.append([t, t[1:-1]])
                        sent.extend([[code[i], nl[j]] for i in xrange(len(code)) for j in xrange(len(nl))])
                        #print [zip(x, code) for x in itertools.permutations(code, len(nl))]
                        #sent.extend([zip(x, code) for x in itertools.permutations(code, len(nl))])
        return sent

    def get_stackoverflow_data_sentences(self, directory):
        sent = []
        for f_path in os.listdir(directory):
            with open(directory+f_path, 'r') as content_file:
                for line in content_file:
                    tokens = line.strip().split(",")
                    if not tokens:
                        continue
                    code = [s for s in tokens if "@" in s]

                    nl = [s for s in tokens if "@" not in s]

                    sent.append(tokens)
                    for t in code:
                        if d.check(t[1:-1]):
                            sent.append([t, t[1:-1]])
                    sent.extend([[code[i], nl[j]] for i in xrange(len(code)) for j in xrange(len(nl))])
                    #print [zip(x, code) for x in itertools.permutations(code, len(nl))]
                    #sent.extend([zip(x, code) for x in itertools.permutations(code, len(nl))])
        return sent

    def get_stackoverflow_data_document(self, directory):
        all_sent = []
        for f_path in os.listdir(directory):
            with open(directory+f_path, 'r') as content_file:
                sent = []
                data=content_file.read().replace('\n', '')

                tokens = data.strip().split(",")
                if not tokens:
                    continue
                document_code = [s for s in tokens if "@" in s]
                document_nl = [s for s in tokens if "@" not in s]

                all_sent.extend([[document_code[i], document_nl[j]] for i in xrange(len(document_code)) for j in xrange(len(document_nl))])
                all_sent.append(tokens)
                for t in code:
                    if d.check(t[1:-1]):
                        sent.append([t, t[1:-1]])
        return all_sent

    #
    # def process_bug_report_data(self, file_name, out_file, project, gitpath, newDir):
    #     wb = load_workbook(filename=file_name)
    #     sheetname = project.lower()
    #     ws = wb[sheetname]
    #     #header = [cell.value for cell in wb.rows[0]]
    #
    #     reports = []
    #
    #     for row in ws.rows[1:]:
    #         args = [cell.value for cell in row]
    #         report = BugReport(*args)
    #         reports.append(report)
    #
    #     os.chdir(gitpath)
    #     first_commit = str(reports[0].commit)
    #     first_report_id = str(reports[0].reportID)
    #
    #     os.system("git checkout " + first_commit+"~1")
    #     newDirPath = newDir + first_report_id + "/"
    #     os.makedirs(newDirPath)
    #     os.system("cp -r * " + newDir)
    #     last_commit = first_commit + "~1"
    #
    #
    #     #because reports is an ordered list, as inserted, so we're inserting as they are in the dataset.
    #     #really have to go and verify this when you build your reconstruction function
    #
    #     for report in reports[1:]:
    #         new_commit = str(report.commit) + "~1"
    #         new_report_id = str(report.reportID)
    #         newDirPath = newDir + new_report_id + "/"
    #         os.makedirs(newDirPath)
    #         os.system("git checkout " + new_commit)
    #         print ('git diff --name-status %s %s | grep ".java$" | grep "^A"' %(last_commit, new_commit))
    #
    #         #print os.system('git diff --name-status %s %s | grep ".java$" | grep "^A"' %(last_commit, new_commit))
    #         #ok I guess I  now only care about any file not about sorting them eh
    #
    #         os.system('git diff --name-status %s %s | grep ".java$" | grep "^A" | cut -f2 | xargs -I "{}" cp {} %s' %(last_commit, new_commit, newDirPath))
    #         os.system('git diff --name-status %s %s | grep ".java$" | grep "^M" | cut -f2 | xargs -I "{}" cp {} %s' %(last_commit, new_commit, newDirPath))
    #         os.system('git diff --name-status %s %s | grep ".java$" | grep "^D"| cut -f2| xargs -I "{}" touch %sDEL_{}' %(last_commit, new_commit, newDirPath))
    #
    #         last_commit = new_commit
    #
    #
    #     #with open(out_file, "wb") as f:
    #     #    for report in reports:
    #     #        f.write(str(report.reportID) + "," + str(report.commit) + "\n")
    #
    #     reportOutFile = project + "_reports_processed.pkl"
    #     pickle.dump(reports, open("/home/ndg/users/carmst16/EmbeddingBugs/resources/bugreport/" + reportOutFile))
    #
    #     return reports



    def process_description(self, text):
        pp = Preprocessor()
        return list(itertools.chain.from_iterable(pp.preprocessLang(text))) #update to return a singe list

    #THIS HAS BEEN CHECKED
    def read_and_process_report_data(self, bug_file_path, project):
        wb = load_workbook(filename=bug_file_path)
        sheetname = project.lower()
        ws = wb[project]
        # header = [cell.value for cell in wb.rows[0]]

        reports = []
        count = 0
        rows = [row for row in ws.rows]
        for row in rows[1:]:
            if count >50:
                break
            count += 1
            args = [cell.value for cell in row]
            report = BugReport(*args)
            reports.append(report)

        for report in reports:
            if not report.description:
                report.description = ""
            text = report.summary + report.description
            report.processed_description = self.process_description(text) #where this is a list of lists of tokenized sentences
            report.files = report.files.replace("java ", "java;").split(";") #NOTE THAT THESE ARE UNICODE. LEAVING FOR NOW

        return reports



    # please note that this no longer outputs the code from a source code file.... only the natural language
    # outputs to the same file structure, but with a different root directory. .txt files not .java
    def process_file(self, infile_path, outfile_path, code=True):
        pp = Preprocessor()
        with open(infile_path, "rb") as f:
            current_comment = False
            current_code = False
            cur_text = ""
            cur_code = ""
            all_tokens = []
            for line  in f:
                line = line.strip()
                if line.startswith("//"):
                    if current_code:
                        if code:
                            tokens = pp.preprocessCode(cur_code)
                            all_tokens.append(tokens) #let's just give this a shot
                        current_code = False
                        cur_code = ""
                    tokens = pp.preprocessLang(line)
                    all_tokens.extend(tokens)
                elif current_comment:
                    cur_text += line
                    if line.endswith("*/"):
                     #   print line, "ended comment"
                        current_comment = False
                        tokens = pp.preprocessLang(cur_text)
                        all_tokens.extend(tokens)
                        cur_text = ""
                    continue
                elif line.startswith("/*"):
                    if current_code:
                        if code:
                            tokens = pp.preprocessCode(cur_code)
                            all_tokens.append(tokens)
                        current_code = False
                        cur_code = ""
                    current_comment = True
                    cur_text += line
                else:
                    if current_code and code:
                        cur_code += line
                    elif code:
                        current_code = True
                        cur_code += line
                    else:
                        pass
        with open(outfile_path, "wb") as outf:
            for l in all_tokens:
                if not l:
                    continue
                s = ",".join(l)
                outf.write(s + "\n")

    #hacky for multiprocess module
    def insert_to_temp(self, report_data):

            report, all_processed_path, base_commit, base_path, temp_path, cloned_repo_path  = report_data
            print report.reportID
            try:
                shutil.rmtree(temp_path, ignore_errors=True)
            except OSError, e:
                 print ("Error: %s - %s." % (e.filename, e.strerror))
            try:
                os.makedirs(temp_path)
            except:
                pass

            #next_path = all_processed_path + str(report.reportID +1 ) + "/"
            #if os.path.exists(next_path): #aka don't bother to do it if it's already been done. requires some maintenance though so be careful
            #    return #todo: check the None doesn't fuck anything up


            #while(os.path.isfile(cloned_repo_path + ".git/index.lock")):
            #    print "sleeping, with report id: ", report.reportID
            #    time.sleep(120)

            prev_commit = base_commit + "~1"
            prev_current_commit = report.commit + "~1"
            os.system("git checkout " + prev_current_commit)
            outfile_path = all_processed_path + str(report.reportID) + "/"

            try:
                os.makedirs(outfile_path)
            except:
                pass
            #print "cp -r %s* %s" %(base_path, outfile_path)
            os.system("\cp -r %s* %s" %(base_path, outfile_path)) #I guess this is the biggest thing to check.....

            #TODO: clean temp
            #temp path but you want to preserve their original file path.....
            # need to check what git diff outputs
            #my working directory should be the cloned_repo_path....
            print "working directory", os.getcwd()
            print 'git diff --name-status %s %s | grep ".java$" | grep "^A" | cut -f2 | xargs -I "{}" cp --parents "{}" %s' % (prev_commit, prev_current_commit, temp_path)
            print 'git diff --name-status %s %s | grep ".java$" | grep "^M" | cut -f2 | xargs -I "{}" cp --parents "{}" %s' % (prev_commit, prev_current_commit, temp_path)
            print 'git diff --name-status %s %s | grep ".java$" | grep "^D"| cut -f2| xargs -I "{}" rm "%s{}"' % (prev_commit, prev_current_commit, outfile_path)
            os.system('git diff --name-status %s %s | grep ".java$" | grep "^A" | cut -f2 | xargs -I "{}" cp --parents "{}" %s' % (prev_commit, prev_current_commit, temp_path))
            os.system('git diff --name-status %s %s | grep ".java$" | grep "^M" | cut -f2 | xargs -I "{}" cp --parents "{}" %s' % (prev_commit, prev_current_commit, temp_path))
            os.system('git diff --name-status %s %s | grep ".java$" | grep "^D"| cut -f2| xargs -I "{}" rm "%s{}.txt"' % (prev_commit, prev_current_commit, outfile_path))

            print "ran copy"


    def temp_to_processed(self, report_data_short):
        temp_path, all_processed_path, id = report_data_short
        print "new processing: ", id
        for dir_, _, files in os.walk(temp_path):
            for fileName in files:
                if not fileName.endswith(".java"):
                    continue
                relDir = os.path.relpath(dir_, temp_path)
                relFile = os.path.join(relDir, fileName)
                infile_path = temp_path + relFile
                outfile_path = all_processed_path + str(id) + "/" + relFile
                to_create = os.path.dirname(outfile_path)
                try:
                    os.makedirs(to_create)
                except:
                    pass
                out_file = outfile_path + ".txt"

                self.process_file(infile_path, out_file)
                os.remove(infile_path)
        print "finished processing: ", id


    def process_all_files(self, cloned_repo_path, reports, all_processed_path, temp_path):
        os.chdir(cloned_repo_path)
        first_report = reports[0]
        base_commit = str(first_report.commit)
        prev_current_commit = base_commit + "~1"
        print "calling_checkout"
        os.system("git checkout " + prev_current_commit)
        print "done"
        base_path = all_processed_path + str(first_report.reportID) + "/"
        print base_path
        count = 0
        for dir_, _, files in os.walk(cloned_repo_path):
            for fileName in files:
                count +=1
                print count
                if not fileName.endswith(".java"):
                    continue
                relDir = os.path.relpath(dir_, cloned_repo_path)
                relFile = os.path.join(relDir, fileName)
                infile_path = cloned_repo_path + relFile
                outfile_path = base_path + relFile
                to_create = os.path.dirname(outfile_path)

                try:
                    os.makedirs(to_create)
                except:
                    continue
                out_file = outfile_path + ".txt"
                if not os.path.isfile(out_file): #so I don't have to reinitialize every time, but I still run a quick sanity check
                    self.process_file(infile_path, out_file)

        report_datas = [(report, all_processed_path, base_commit, base_path, temp_path+str(report.reportID)+"/", cloned_repo_path) for report in reports[14:15]]
        report_datas_short = [(temp_path+str(report.reportID)+"/", all_processed_path, report.reportID) for report in reports[14:15]] #this is so super dumb but I have no more time and patience sorry future caitrin

        for report_data in report_datas:
            self.insert_to_temp(report_data)

        #after you've finished copying to the right temp folder, blast through the processing.
        pool = pp.ProcessPool(16)
        res = pool.map(self.temp_to_processed, report_datas_short)
        pool.close()
        pool.join()


    # where datapath is the freshly cloned repocd ..
    def create_file_repo(self, data_path, report, processed_path):
        os.chdir(data_path)

        first_commit = str(report.commit)
        prev_current_commit = first_commit + "~1"
        os.system("git checkout " + prev_current_commit)

        fileSet = set()

        for dir_, _, files in os.walk(data_path):

            for fileName in files:
                if not fileName.endswith(".java"):
                    continue
                relDir = os.path.relpath(dir_, data_path)
                relFile = os.path.join(relDir, fileName)
                infile_path = data_path + relFile
                outfile_path = processed_path + relFile
                to_create = os.path.dirname(outfile_path)
                if not os.path.exists(to_create):
                    os.makedirs(to_create)
                out_file = outfile_path + ".txt"
                self.process_file(infile_path, out_file)



    # where these are the raw from the sheet, unprocessed
    def ahhhhh(self, previous_commit, current_commit, data_path, temp_path, processed_path):
        os.system("git checkout " + current_commit)
        prev_last_commit = previous_commit + "~1"
        prev_current_commit = current_commit + "~1"

        #temp path but you want to preserve their original file path.....
        # need to check what git diff outputs

        #os.system(
        #    'git diff --name-status %s %s | grep ".java$" | grep "^A" | cut -f2 | xargs -I "{}" cp --parents "{}" %s' % (
        #    prev_last_commit, prev_current_commit, temp_path))
        #os.system(
        #    'git diff --name-status %s %s | grep ".java$" | grep "^M" | cut -f2 | xargs -I "{}" cp --parents "{}" %s' % (
        #    prev_last_commit, prev_current_commit, temp_path))
        os.system('git diff --name-status %s %s | grep ".java$" | grep "^D"| cut -f2| xargs -I "{}" rm "%s{}"' % (
        prev_last_commit, prev_current_commit, processed_path))



def readBugReport():
    bug_reports = []
    total = 0
    i=0
    path = "dataset/"
    #project_files = ['Birt.xml', 'Eclipse_Platform_UI.xml', 'JDT.xml', 'SWT.xml']
    project_files = ['Birt.xml', 'Eclipse_Platform_UI.xml']
    for file in project_files:
        tree = ET.parse(path+file)
        root = tree.getroot()
        for table in root.iter('table'):
            i=i+1
            for column in table.findall('column'):
                if column is not None:
                    name = column.get('name')
                    if name == "id":
                        reportID = column.text
                        #print(reportID)
                    if name=="bug_id":
                        bug_id = column.text
                        #print(bug_id)
                    if name == "summary":
                        summary = column.text
                        #print(summary)
                    if name=="description":
                        description = column.text
                        #print(description)
                    if name == "report_time":
                        report_time = column.text
                        #print(report_time)
                    if name=="report_timestamp":
                        report_timestamp = column.text
                        #print(report_timestamp)
                    if name == "status":
                        status = column.text
                        #print(status)
                    if name=="commit":
                        commit = column.text
                        #print(commit)
                    if name == "commit_timestamp":
                        commit_timestamp = column.text
                        #print(commit_timestamp)
                    if name=="files":
                        files = column.text
                        #print(files)

            #1.Bug reports with status marked as resolved fixed, veri- fied fixed, or closed fixed were collected for evaluation
            if (status=="resolved fixed" or status =="closed fixed" or status=="verified fixed"):
                ## 3.  Bug reports without fixed files are also ignored because they are considered not functional
                if (files is not None):
                    total= total+1
                    bug_reports.append(BugReport(reportID,bug_id,summary,description,report_time,report_timestamp,status,commit,commit_timestamp,files))
    print("all Total:"+str(i))
    print("final report Total:" + str(total))
    return(bug_reports)


#bug_reports = readBugReport()
##### This is just for reference, how you can access each bug report !!!
#for bug_report in bug_reports:
    #print(bug_report.bug_id +" "+bug_report.summary)
